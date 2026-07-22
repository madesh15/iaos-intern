"""Logistics & Freight — Internal Audit module.
Mounted at /api/modules/logistics_freight.
"""
import csv
import io
from datetime import date, datetime
from typing import Optional

from fastapi import APIRouter, HTTPException, Query, Response
from sqlalchemy import or_

from app.api.deps import CurrentUser, DbSession
from app.core.tenancy import tenant_scoped

from . import analytics as an
from . import services as svc
from .models import (
    Carrier, Vehicle, Route, FreightRateContract, FreightShipment,
    FreightInvoice, POD, FuelIndex, Claim, DetentionCharge,
    DashboardKPI, Finding, ActionTracker,
    Plant, Warehouse, Region, BusinessUnit,
    RiskControl, TestRule, DataSource, SamplingRecord,
    ExceptionItem, WorkingPaper,
)
from .schemas import (
    CarrierCreate, CarrierUpdate, CarrierOut,
    VehicleCreate, VehicleUpdate, VehicleOut,
    RouteCreate, RouteUpdate, RouteOut,
    FreightRateContractCreate, FreightRateContractUpdate, FreightRateContractOut,
    FreightShipmentCreate, FreightShipmentUpdate, FreightShipmentOut,
    FreightInvoiceCreate, FreightInvoiceUpdate, FreightInvoiceOut,
    PODCreate, PODUpdate, PODOut,
    FuelIndexCreate, FuelIndexUpdate, FuelIndexOut,
    ClaimCreate, ClaimUpdate, ClaimOut,
    DetentionChargeCreate, DetentionChargeUpdate, DetentionChargeOut,
    DashboardKPICreate, DashboardKPIOut,
    FindingCreate, FindingUpdate, FindingOut,
    ActionTrackerCreate, ActionTrackerUpdate, ActionTrackerOut,
    PlantCreate, PlantUpdate, PlantOut,
    WarehouseCreate, WarehouseUpdate, WarehouseOut,
    RegionCreate, RegionUpdate, RegionOut,
    BusinessUnitCreate, BusinessUnitUpdate, BusinessUnitOut,
    RiskControlCreate, RiskControlUpdate, RiskControlOut,
    TestRuleCreate, TestRuleUpdate, TestRuleOut,
    DataSourceCreate, DataSourceUpdate, DataSourceOut,
    SamplingRecordCreate, SamplingRecordUpdate, SamplingRecordOut,
    ExceptionItemCreate, ExceptionItemUpdate, ExceptionItemOut,
    WorkingPaperCreate, WorkingPaperUpdate, WorkingPaperOut,
)

MANIFEST = {
    "name": "logistics_freight",
    "title": "Logistics & Freight",
    "description": "Controls freight and transport costs through freight rate validation, routing analytics, carrier performance, detention analysis, duplicate billing detection, POD reconciliation, SLA monitoring, and freight cost analytics.",
    "icon": "truck",
    "group": "Operations",
    "industry": "",
    "version": "1.0.0",
    "owner": "audit",
}

router = APIRouter()


def _q(db, model, current_user):
    """Tenant-scoped, non-deleted query helper."""
    return tenant_scoped(
        db.query(model).filter(model.is_deleted == False), current_user
    )


def _paginate(query, page: int, page_size: int):
    total = query.count()
    items = query.offset((page - 1) * page_size).limit(page_size).all()
    return {
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": max(1, (total + page_size - 1) // page_size),
    }


# ──────────────────────────────────────────────
# CARRIER CRUD
# ──────────────────────────────────────────────

@router.get("/carriers", response_model=dict)
def list_carriers(
    db: DbSession, current_user: CurrentUser,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    search: str = "",
    status: str = "",
):
    q = _q(db, Carrier, current_user)
    if search:
        q = q.filter(Carrier.name.ilike(f"%{search}%") | Carrier.code.ilike(f"%{search}%"))
    if status:
        q = q.filter(Carrier.status == status)
    q = q.order_by(Carrier.id.desc())
    result = _paginate(q, page, page_size)
    result["items"] = [CarrierOut.model_validate(i) for i in result["items"]]
    return result


@router.get("/carriers/{carrier_id}", response_model=CarrierOut)
def get_carrier(carrier_id: int, db: DbSession, current_user: CurrentUser):
    c = _q(db, Carrier, current_user).filter(Carrier.id == carrier_id).first()
    if not c:
        raise HTTPException(404, "Carrier not found")
    return CarrierOut.model_validate(c)


@router.post("/carriers", response_model=CarrierOut, status_code=201)
def create_carrier(body: CarrierCreate, db: DbSession, current_user: CurrentUser):
    existing = _q(db, Carrier, current_user).filter(Carrier.code == body.code).first()
    if existing:
        raise HTTPException(409, "Carrier code already exists")
    c = Carrier(**body.model_dump(), tenant_id=current_user.tenant_id)
    db.add(c)
    db.commit()
    db.refresh(c)
    return CarrierOut.model_validate(c)


@router.put("/carriers/{carrier_id}", response_model=CarrierOut)
def update_carrier(carrier_id: int, body: CarrierUpdate, db: DbSession, current_user: CurrentUser):
    c = _q(db, Carrier, current_user).filter(Carrier.id == carrier_id).first()
    if not c:
        raise HTTPException(404, "Carrier not found")
    for k, v in body.model_dump(exclude_unset=True).items():
        setattr(c, k, v)
    db.commit()
    db.refresh(c)
    return CarrierOut.model_validate(c)


@router.delete("/carriers/{carrier_id}", status_code=204)
def delete_carrier(carrier_id: int, db: DbSession, current_user: CurrentUser):
    c = _q(db, Carrier, current_user).filter(Carrier.id == carrier_id).first()
    if not c:
        raise HTTPException(404, "Carrier not found")
    c.is_deleted = True
    db.commit()


# ──────────────────────────────────────────────
# VEHICLE CRUD
# ──────────────────────────────────────────────

@router.get("/vehicles", response_model=dict)
def list_vehicles(
    db: DbSession, current_user: CurrentUser,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    search: str = "",
    carrier_id: int = 0,
):
    q = _q(db, Vehicle, current_user)
    if search:
        q = q.filter(Vehicle.registration_number.ilike(f"%{search}%"))
    if carrier_id:
        q = q.filter(Vehicle.carrier_id == carrier_id)
    q = q.order_by(Vehicle.id.desc())
    result = _paginate(q, page, page_size)
    result["items"] = [VehicleOut.model_validate(i) for i in result["items"]]
    return result


@router.get("/vehicles/{vehicle_id}", response_model=VehicleOut)
def get_vehicle(vehicle_id: int, db: DbSession, current_user: CurrentUser):
    v = _q(db, Vehicle, current_user).filter(Vehicle.id == vehicle_id).first()
    if not v:
        raise HTTPException(404, "Vehicle not found")
    return VehicleOut.model_validate(v)


@router.post("/vehicles", response_model=VehicleOut, status_code=201)
def create_vehicle(body: VehicleCreate, db: DbSession, current_user: CurrentUser):
    carrier = _q(db, Carrier, current_user).filter(Carrier.id == body.carrier_id).first()
    if not carrier:
        raise HTTPException(404, "Carrier not found")
    v = Vehicle(**body.model_dump(), tenant_id=current_user.tenant_id)
    db.add(v)
    db.commit()
    db.refresh(v)
    return VehicleOut.model_validate(v)


@router.put("/vehicles/{vehicle_id}", response_model=VehicleOut)
def update_vehicle(vehicle_id: int, body: VehicleUpdate, db: DbSession, current_user: CurrentUser):
    v = _q(db, Vehicle, current_user).filter(Vehicle.id == vehicle_id).first()
    if not v:
        raise HTTPException(404, "Vehicle not found")
    for k, val in body.model_dump(exclude_unset=True).items():
        setattr(v, k, val)
    db.commit()
    db.refresh(v)
    return VehicleOut.model_validate(v)


@router.delete("/vehicles/{vehicle_id}", status_code=204)
def delete_vehicle(vehicle_id: int, db: DbSession, current_user: CurrentUser):
    v = _q(db, Vehicle, current_user).filter(Vehicle.id == vehicle_id).first()
    if not v:
        raise HTTPException(404, "Vehicle not found")
    v.is_deleted = True
    db.commit()


# ──────────────────────────────────────────────
# ROUTE CRUD
# ──────────────────────────────────────────────

@router.get("/routes", response_model=dict)
def list_routes(
    db: DbSession, current_user: CurrentUser,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    search: str = "",
    mode: str = "",
):
    q = _q(db, Route, current_user)
    if search:
        q = q.filter(Route.origin.ilike(f"%{search}%") | Route.destination.ilike(f"%{search}%"))
    if mode:
        q = q.filter(Route.mode == mode)
    q = q.order_by(Route.id.desc())
    result = _paginate(q, page, page_size)
    result["items"] = [RouteOut.model_validate(i) for i in result["items"]]
    return result


@router.get("/routes/{route_id}", response_model=RouteOut)
def get_route(route_id: int, db: DbSession, current_user: CurrentUser):
    r = _q(db, Route, current_user).filter(Route.id == route_id).first()
    if not r:
        raise HTTPException(404, "Route not found")
    return RouteOut.model_validate(r)


@router.post("/routes", response_model=RouteOut, status_code=201)
def create_route(body: RouteCreate, db: DbSession, current_user: CurrentUser):
    r = Route(**body.model_dump(), tenant_id=current_user.tenant_id)
    db.add(r)
    db.commit()
    db.refresh(r)
    return RouteOut.model_validate(r)


@router.put("/routes/{route_id}", response_model=RouteOut)
def update_route(route_id: int, body: RouteUpdate, db: DbSession, current_user: CurrentUser):
    r = _q(db, Route, current_user).filter(Route.id == route_id).first()
    if not r:
        raise HTTPException(404, "Route not found")
    for k, val in body.model_dump(exclude_unset=True).items():
        setattr(r, k, val)
    db.commit()
    db.refresh(r)
    return RouteOut.model_validate(r)


@router.delete("/routes/{route_id}", status_code=204)
def delete_route(route_id: int, db: DbSession, current_user: CurrentUser):
    r = _q(db, Route, current_user).filter(Route.id == route_id).first()
    if not r:
        raise HTTPException(404, "Route not found")
    r.is_deleted = True
    db.commit()


# ──────────────────────────────────────────────
# FREIGHT RATE CONTRACT CRUD
# ──────────────────────────────────────────────

@router.get("/contracts", response_model=dict)
def list_contracts(
    db: DbSession, current_user: CurrentUser,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    search: str = "",
    status: str = "",
):
    q = _q(db, FreightRateContract, current_user)
    if search:
        q = q.filter(FreightRateContract.contract_number.ilike(f"%{search}%"))
    if status:
        q = q.filter(FreightRateContract.status == status)
    q = q.order_by(FreightRateContract.id.desc())
    result = _paginate(q, page, page_size)
    result["items"] = [FreightRateContractOut.model_validate(i) for i in result["items"]]
    return result


@router.get("/contracts/{contract_id}", response_model=FreightRateContractOut)
def get_contract(contract_id: int, db: DbSession, current_user: CurrentUser):
    c = _q(db, FreightRateContract, current_user).filter(FreightRateContract.id == contract_id).first()
    if not c:
        raise HTTPException(404, "Contract not found")
    return FreightRateContractOut.model_validate(c)


@router.post("/contracts", response_model=FreightRateContractOut, status_code=201)
def create_contract(body: FreightRateContractCreate, db: DbSession, current_user: CurrentUser):
    carrier = _q(db, Carrier, current_user).filter(Carrier.id == body.carrier_id).first()
    if not carrier:
        raise HTTPException(404, "Carrier not found")
    c = FreightRateContract(**body.model_dump(), tenant_id=current_user.tenant_id)
    db.add(c)
    db.commit()
    db.refresh(c)
    return FreightRateContractOut.model_validate(c)


@router.put("/contracts/{contract_id}", response_model=FreightRateContractOut)
def update_contract(contract_id: int, body: FreightRateContractUpdate, db: DbSession, current_user: CurrentUser):
    c = _q(db, FreightRateContract, current_user).filter(FreightRateContract.id == contract_id).first()
    if not c:
        raise HTTPException(404, "Contract not found")
    for k, val in body.model_dump(exclude_unset=True).items():
        setattr(c, k, val)
    db.commit()
    db.refresh(c)
    return FreightRateContractOut.model_validate(c)


@router.delete("/contracts/{contract_id}", status_code=204)
def delete_contract(contract_id: int, db: DbSession, current_user: CurrentUser):
    c = _q(db, FreightRateContract, current_user).filter(FreightRateContract.id == contract_id).first()
    if not c:
        raise HTTPException(404, "Contract not found")
    c.is_deleted = True
    db.commit()


# ──────────────────────────────────────────────
# FREIGHT SHIPMENT CRUD
# ──────────────────────────────────────────────

@router.get("/shipments", response_model=dict)
def list_shipments(
    db: DbSession, current_user: CurrentUser,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    search: str = "",
    status: str = "",
    carrier_id: int = 0,
    mode: str = "",
    date_from: str = "",
    date_to: str = "",
):
    q = _q(db, FreightShipment, current_user)
    if search:
        q = q.filter(
            FreightShipment.shipment_number.ilike(f"%{search}%")
            | FreightShipment.lr_number.ilike(f"%{search}%")
            | FreightShipment.origin.ilike(f"%{search}%")
            | FreightShipment.destination.ilike(f"%{search}%")
        )
    if status:
        q = q.filter(FreightShipment.status == status)
    if carrier_id:
        q = q.filter(FreightShipment.carrier_id == carrier_id)
    if mode:
        q = q.filter(FreightShipment.mode == mode)
    if date_from:
        q = q.filter(FreightShipment.shipment_date >= date.fromisoformat(date_from))
    if date_to:
        q = q.filter(FreightShipment.shipment_date <= date.fromisoformat(date_to))
    q = q.order_by(FreightShipment.id.desc())
    result = _paginate(q, page, page_size)
    result["items"] = [FreightShipmentOut.model_validate(i) for i in result["items"]]
    return result


@router.get("/shipments/{shipment_id}", response_model=FreightShipmentOut)
def get_shipment(shipment_id: int, db: DbSession, current_user: CurrentUser):
    s = _q(db, FreightShipment, current_user).filter(FreightShipment.id == shipment_id).first()
    if not s:
        raise HTTPException(404, "Shipment not found")
    return FreightShipmentOut.model_validate(s)


@router.post("/shipments", response_model=FreightShipmentOut, status_code=201)
def create_shipment(body: FreightShipmentCreate, db: DbSession, current_user: CurrentUser):
    carrier = _q(db, Carrier, current_user).filter(Carrier.id == body.carrier_id).first()
    if not carrier:
        raise HTTPException(404, "Carrier not found")
    existing = _q(db, FreightShipment, current_user).filter(FreightShipment.shipment_number == body.shipment_number).first()
    if existing:
        raise HTTPException(409, "Shipment number already exists")
    if body.lr_number:
        lr_exists = _q(db, FreightShipment, current_user).filter(FreightShipment.lr_number == body.lr_number).first()
        if lr_exists:
            raise HTTPException(409, "LR number already exists")
    if body.total_amount < 0:
        raise HTTPException(422, "Freight amount cannot be negative")
    if body.shipment_date > body.expected_delivery_date:
        raise HTTPException(422, "Shipment date cannot be after expected delivery date")
    s = FreightShipment(**body.model_dump(), tenant_id=current_user.tenant_id)
    db.add(s)
    db.commit()
    db.refresh(s)
    return FreightShipmentOut.model_validate(s)


@router.put("/shipments/{shipment_id}", response_model=FreightShipmentOut)
def update_shipment(shipment_id: int, body: FreightShipmentUpdate, db: DbSession, current_user: CurrentUser):
    s = _q(db, FreightShipment, current_user).filter(FreightShipment.id == shipment_id).first()
    if not s:
        raise HTTPException(404, "Shipment not found")
    for k, val in body.model_dump(exclude_unset=True).items():
        setattr(s, k, val)
    db.commit()
    db.refresh(s)
    return FreightShipmentOut.model_validate(s)


@router.delete("/shipments/{shipment_id}", status_code=204)
def delete_shipment(shipment_id: int, db: DbSession, current_user: CurrentUser):
    s = _q(db, FreightShipment, current_user).filter(FreightShipment.id == shipment_id).first()
    if not s:
        raise HTTPException(404, "Shipment not found")
    s.is_deleted = True
    db.commit()


# ──────────────────────────────────────────────
# FREIGHT INVOICE CRUD
# ──────────────────────────────────────────────

@router.get("/invoices", response_model=dict)
def list_invoices(
    db: DbSession, current_user: CurrentUser,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    search: str = "",
    status: str = "",
    payment_status: str = "",
):
    q = _q(db, FreightInvoice, current_user)
    if search:
        q = q.filter(FreightInvoice.invoice_number.ilike(f"%{search}%"))
    if status:
        q = q.filter(FreightInvoice.status == status)
    if payment_status:
        q = q.filter(FreightInvoice.payment_status == payment_status)
    q = q.order_by(FreightInvoice.id.desc())
    result = _paginate(q, page, page_size)
    result["items"] = [FreightInvoiceOut.model_validate(i) for i in result["items"]]
    return result


@router.get("/invoices/{invoice_id}", response_model=FreightInvoiceOut)
def get_invoice(invoice_id: int, db: DbSession, current_user: CurrentUser):
    i = _q(db, FreightInvoice, current_user).filter(FreightInvoice.id == invoice_id).first()
    if not i:
        raise HTTPException(404, "Invoice not found")
    return FreightInvoiceOut.model_validate(i)


@router.post("/invoices", response_model=FreightInvoiceOut, status_code=201)
def create_invoice(body: FreightInvoiceCreate, db: DbSession, current_user: CurrentUser):
    carrier = _q(db, Carrier, current_user).filter(Carrier.id == body.carrier_id).first()
    if not carrier:
        raise HTTPException(404, "Carrier not found")
    inv_exists = _q(db, FreightInvoice, current_user).filter(FreightInvoice.invoice_number == body.invoice_number).first()
    if inv_exists:
        raise HTTPException(409, "Invoice number already exists")
    if body.total_amount < 0:
        raise HTTPException(422, "Invoice amount cannot be negative")
    if body.due_date < body.invoice_date:
        raise HTTPException(422, "Due date cannot be before invoice date")
    i = FreightInvoice(**body.model_dump(), tenant_id=current_user.tenant_id)
    db.add(i)
    db.commit()
    db.refresh(i)
    return FreightInvoiceOut.model_validate(i)


@router.put("/invoices/{invoice_id}", response_model=FreightInvoiceOut)
def update_invoice(invoice_id: int, body: FreightInvoiceUpdate, db: DbSession, current_user: CurrentUser):
    i = _q(db, FreightInvoice, current_user).filter(FreightInvoice.id == invoice_id).first()
    if not i:
        raise HTTPException(404, "Invoice not found")
    for k, val in body.model_dump(exclude_unset=True).items():
        setattr(i, k, val)
    db.commit()
    db.refresh(i)
    return FreightInvoiceOut.model_validate(i)


@router.delete("/invoices/{invoice_id}", status_code=204)
def delete_invoice(invoice_id: int, db: DbSession, current_user: CurrentUser):
    i = _q(db, FreightInvoice, current_user).filter(FreightInvoice.id == invoice_id).first()
    if not i:
        raise HTTPException(404, "Invoice not found")
    i.is_deleted = True
    db.commit()


# ──────────────────────────────────────────────
# POD CRUD
# ──────────────────────────────────────────────

@router.get("/pods", response_model=dict)
def list_pods(
    db: DbSession, current_user: CurrentUser,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    search: str = "",
):
    q = _q(db, POD, current_user)
    if search:
        q = q.filter(POD.pod_number.ilike(f"%{search}%"))
    q = q.order_by(POD.id.desc())
    result = _paginate(q, page, page_size)
    result["items"] = [PODOut.model_validate(i) for i in result["items"]]
    return result


@router.get("/pods/{pod_id}", response_model=PODOut)
def get_pod(pod_id: int, db: DbSession, current_user: CurrentUser):
    p = _q(db, POD, current_user).filter(POD.id == pod_id).first()
    if not p:
        raise HTTPException(404, "POD not found")
    return PODOut.model_validate(p)


@router.post("/pods", response_model=PODOut, status_code=201)
def create_pod(body: PODCreate, db: DbSession, current_user: CurrentUser):
    shipment = _q(db, FreightShipment, current_user).filter(FreightShipment.id == body.shipment_id).first()
    if not shipment:
        raise HTTPException(404, "Shipment not found")
    existing_pod = _q(db, POD, current_user).filter(POD.shipment_id == body.shipment_id).first()
    if existing_pod:
        raise HTTPException(409, "POD already exists for this shipment")
    p = POD(**body.model_dump(), tenant_id=current_user.tenant_id)
    db.add(p)
    db.commit()
    db.refresh(p)
    return PODOut.model_validate(p)


@router.put("/pods/{pod_id}", response_model=PODOut)
def update_pod(pod_id: int, body: PODUpdate, db: DbSession, current_user: CurrentUser):
    p = _q(db, POD, current_user).filter(POD.id == pod_id).first()
    if not p:
        raise HTTPException(404, "POD not found")
    for k, val in body.model_dump(exclude_unset=True).items():
        setattr(p, k, val)
    db.commit()
    db.refresh(p)
    return PODOut.model_validate(p)


@router.delete("/pods/{pod_id}", status_code=204)
def delete_pod(pod_id: int, db: DbSession, current_user: CurrentUser):
    p = _q(db, POD, current_user).filter(POD.id == pod_id).first()
    if not p:
        raise HTTPException(404, "POD not found")
    p.is_deleted = True
    db.commit()


# ──────────────────────────────────────────────
# FUEL INDEX CRUD
# ──────────────────────────────────────────────

@router.get("/fuel-indices", response_model=dict)
def list_fuel_indices(
    db: DbSession, current_user: CurrentUser,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
):
    q = _q(db, FuelIndex, current_user).order_by(FuelIndex.index_date.desc())
    result = _paginate(q, page, page_size)
    result["items"] = [FuelIndexOut.model_validate(i) for i in result["items"]]
    return result


@router.get("/fuel-indices/{fuel_id}", response_model=FuelIndexOut)
def get_fuel_index(fuel_id: int, db: DbSession, current_user: CurrentUser):
    f = _q(db, FuelIndex, current_user).filter(FuelIndex.id == fuel_id).first()
    if not f:
        raise HTTPException(404, "Fuel index not found")
    return FuelIndexOut.model_validate(f)


@router.post("/fuel-indices", response_model=FuelIndexOut, status_code=201)
def create_fuel_index(body: FuelIndexCreate, db: DbSession, current_user: CurrentUser):
    f = FuelIndex(**body.model_dump(), tenant_id=current_user.tenant_id)
    db.add(f)
    db.commit()
    db.refresh(f)
    return FuelIndexOut.model_validate(f)


@router.put("/fuel-indices/{fuel_id}", response_model=FuelIndexOut)
def update_fuel_index(fuel_id: int, body: FuelIndexUpdate, db: DbSession, current_user: CurrentUser):
    f = _q(db, FuelIndex, current_user).filter(FuelIndex.id == fuel_id).first()
    if not f:
        raise HTTPException(404, "Fuel index not found")
    for k, val in body.model_dump(exclude_unset=True).items():
        setattr(f, k, val)
    db.commit()
    db.refresh(f)
    return FuelIndexOut.model_validate(f)


@router.delete("/fuel-indices/{fuel_id}", status_code=204)
def delete_fuel_index(fuel_id: int, db: DbSession, current_user: CurrentUser):
    f = _q(db, FuelIndex, current_user).filter(FuelIndex.id == fuel_id).first()
    if not f:
        raise HTTPException(404, "Fuel index not found")
    f.is_deleted = True
    db.commit()


# ──────────────────────────────────────────────
# CLAIM CRUD
# ──────────────────────────────────────────────

@router.get("/claims", response_model=dict)
def list_claims(
    db: DbSession, current_user: CurrentUser,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    search: str = "",
    status: str = "",
    claim_type: str = "",
):
    q = _q(db, Claim, current_user)
    if search:
        q = q.filter(Claim.claim_number.ilike(f"%{search}%"))
    if status:
        q = q.filter(Claim.status == status)
    if claim_type:
        q = q.filter(Claim.claim_type == claim_type)
    q = q.order_by(Claim.id.desc())
    result = _paginate(q, page, page_size)
    result["items"] = [ClaimOut.model_validate(i) for i in result["items"]]
    return result


@router.get("/claims/{claim_id}", response_model=ClaimOut)
def get_claim(claim_id: int, db: DbSession, current_user: CurrentUser):
    c = _q(db, Claim, current_user).filter(Claim.id == claim_id).first()
    if not c:
        raise HTTPException(404, "Claim not found")
    return ClaimOut.model_validate(c)


@router.post("/claims", response_model=ClaimOut, status_code=201)
def create_claim(body: ClaimCreate, db: DbSession, current_user: CurrentUser):
    shipment = _q(db, FreightShipment, current_user).filter(FreightShipment.id == body.shipment_id).first()
    if not shipment:
        raise HTTPException(404, "Shipment not found")
    c = Claim(**body.model_dump(), tenant_id=current_user.tenant_id)
    db.add(c)
    db.commit()
    db.refresh(c)
    return ClaimOut.model_validate(c)


@router.put("/claims/{claim_id}", response_model=ClaimOut)
def update_claim(claim_id: int, body: ClaimUpdate, db: DbSession, current_user: CurrentUser):
    c = _q(db, Claim, current_user).filter(Claim.id == claim_id).first()
    if not c:
        raise HTTPException(404, "Claim not found")
    for k, val in body.model_dump(exclude_unset=True).items():
        setattr(c, k, val)
    db.commit()
    db.refresh(c)
    return ClaimOut.model_validate(c)


@router.delete("/claims/{claim_id}", status_code=204)
def delete_claim(claim_id: int, db: DbSession, current_user: CurrentUser):
    c = _q(db, Claim, current_user).filter(Claim.id == claim_id).first()
    if not c:
        raise HTTPException(404, "Claim not found")
    c.is_deleted = True
    db.commit()


# ──────────────────────────────────────────────
# DETENTION CHARGE CRUD
# ──────────────────────────────────────────────

@router.get("/detention-charges", response_model=dict)
def list_detention_charges(
    db: DbSession, current_user: CurrentUser,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    shipment_id: int = 0,
):
    q = _q(db, DetentionCharge, current_user)
    if shipment_id:
        q = q.filter(DetentionCharge.shipment_id == shipment_id)
    q = q.order_by(DetentionCharge.id.desc())
    result = _paginate(q, page, page_size)
    result["items"] = [DetentionChargeOut.model_validate(i) for i in result["items"]]
    return result


@router.get("/detention-charges/{charge_id}", response_model=DetentionChargeOut)
def get_detention_charge(charge_id: int, db: DbSession, current_user: CurrentUser):
    d = _q(db, DetentionCharge, current_user).filter(DetentionCharge.id == charge_id).first()
    if not d:
        raise HTTPException(404, "Detention charge not found")
    return DetentionChargeOut.model_validate(d)


@router.post("/detention-charges", response_model=DetentionChargeOut, status_code=201)
def create_detention_charge(body: DetentionChargeCreate, db: DbSession, current_user: CurrentUser):
    shipment = _q(db, FreightShipment, current_user).filter(FreightShipment.id == body.shipment_id).first()
    if not shipment:
        raise HTTPException(404, "Shipment not found")
    d = DetentionCharge(**body.model_dump(), tenant_id=current_user.tenant_id)
    db.add(d)
    db.commit()
    db.refresh(d)
    return DetentionChargeOut.model_validate(d)


@router.put("/detention-charges/{charge_id}", response_model=DetentionChargeOut)
def update_detention_charge(charge_id: int, body: DetentionChargeUpdate, db: DbSession, current_user: CurrentUser):
    d = _q(db, DetentionCharge, current_user).filter(DetentionCharge.id == charge_id).first()
    if not d:
        raise HTTPException(404, "Detention charge not found")
    for k, val in body.model_dump(exclude_unset=True).items():
        setattr(d, k, val)
    db.commit()
    db.refresh(d)
    return DetentionChargeOut.model_validate(d)


@router.delete("/detention-charges/{charge_id}", status_code=204)
def delete_detention_charge(charge_id: int, db: DbSession, current_user: CurrentUser):
    d = _q(db, DetentionCharge, current_user).filter(DetentionCharge.id == charge_id).first()
    if not d:
        raise HTTPException(404, "Detention charge not found")
    d.is_deleted = True
    db.commit()


# ──────────────────────────────────────────────
# DASHBOARD KPI CRUD
# ──────────────────────────────────────────────

@router.get("/kpis", response_model=dict)
def list_kpis(
    db: DbSession, current_user: CurrentUser,
    category: str = "",
    period: str = "",
):
    q = _q(db, DashboardKPI, current_user)
    if category:
        q = q.filter(DashboardKPI.kpi_category == category)
    if period:
        q = q.filter(DashboardKPI.period == period)
    q = q.order_by(DashboardKPI.period_start.desc())
    items = q.all()
    return {"items": [DashboardKPIOut.model_validate(i) for i in items], "total": len(items)}


@router.post("/kpis", response_model=DashboardKPIOut, status_code=201)
def create_kpi(body: DashboardKPICreate, db: DbSession, current_user: CurrentUser):
    k = DashboardKPI(**body.model_dump(), tenant_id=current_user.tenant_id)
    db.add(k)
    db.commit()
    db.refresh(k)
    return DashboardKPIOut.model_validate(k)


# ──────────────────────────────────────────────
# FINDING CRUD
# ──────────────────────────────────────────────

@router.get("/findings", response_model=dict)
def list_findings(
    db: DbSession, current_user: CurrentUser,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    search: str = "",
    severity: str = "",
    status: str = "",
    category: str = "",
):
    q = _q(db, Finding, current_user)
    if search:
        q = q.filter(
            Finding.finding_number.ilike(f"%{search}%")
            | Finding.title.ilike(f"%{search}%")
        )
    if severity:
        q = q.filter(Finding.severity == severity)
    if status:
        q = q.filter(Finding.status == status)
    if category:
        q = q.filter(Finding.category == category)
    q = q.order_by(Finding.id.desc())
    result = _paginate(q, page, page_size)
    result["items"] = [FindingOut.model_validate(i) for i in result["items"]]
    return result


@router.get("/findings/{finding_id}", response_model=FindingOut)
def get_finding(finding_id: int, db: DbSession, current_user: CurrentUser):
    f = _q(db, Finding, current_user).filter(Finding.id == finding_id).first()
    if not f:
        raise HTTPException(404, "Finding not found")
    return FindingOut.model_validate(f)


@router.post("/findings", response_model=FindingOut, status_code=201)
def create_finding(body: FindingCreate, db: DbSession, current_user: CurrentUser):
    f = Finding(**body.model_dump(), tenant_id=current_user.tenant_id)
    db.add(f)
    db.commit()
    db.refresh(f)
    return FindingOut.model_validate(f)


@router.put("/findings/{finding_id}", response_model=FindingOut)
def update_finding(finding_id: int, body: FindingUpdate, db: DbSession, current_user: CurrentUser):
    f = _q(db, Finding, current_user).filter(Finding.id == finding_id).first()
    if not f:
        raise HTTPException(404, "Finding not found")
    for k, val in body.model_dump(exclude_unset=True).items():
        setattr(f, k, val)
    db.commit()
    db.refresh(f)
    return FindingOut.model_validate(f)


@router.delete("/findings/{finding_id}", status_code=204)
def delete_finding(finding_id: int, db: DbSession, current_user: CurrentUser):
    f = _q(db, Finding, current_user).filter(Finding.id == finding_id).first()
    if not f:
        raise HTTPException(404, "Finding not found")
    f.is_deleted = True
    db.commit()


# ──────────────────────────────────────────────
# ACTION TRACKER CRUD
# ──────────────────────────────────────────────

@router.get("/actions", response_model=dict)
def list_actions(
    db: DbSession, current_user: CurrentUser,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    search: str = "",
    status: str = "",
    priority: str = "",
):
    q = _q(db, ActionTracker, current_user)
    if search:
        q = q.filter(
            ActionTracker.action_number.ilike(f"%{search}%")
            | ActionTracker.title.ilike(f"%{search}%")
            | ActionTracker.assigned_to.ilike(f"%{search}%")
        )
    if status:
        q = q.filter(ActionTracker.status == status)
    if priority:
        q = q.filter(ActionTracker.priority == priority)
    q = q.order_by(ActionTracker.id.desc())
    result = _paginate(q, page, page_size)
    result["items"] = [ActionTrackerOut.model_validate(i) for i in result["items"]]
    return result


@router.get("/actions/{action_id}", response_model=ActionTrackerOut)
def get_action(action_id: int, db: DbSession, current_user: CurrentUser):
    a = _q(db, ActionTracker, current_user).filter(ActionTracker.id == action_id).first()
    if not a:
        raise HTTPException(404, "Action not found")
    return ActionTrackerOut.model_validate(a)


@router.post("/actions", response_model=ActionTrackerOut, status_code=201)
def create_action(body: ActionTrackerCreate, db: DbSession, current_user: CurrentUser):
    a = ActionTracker(**body.model_dump(), tenant_id=current_user.tenant_id)
    db.add(a)
    db.commit()
    db.refresh(a)
    return ActionTrackerOut.model_validate(a)


@router.put("/actions/{action_id}", response_model=ActionTrackerOut)
def update_action(action_id: int, body: ActionTrackerUpdate, db: DbSession, current_user: CurrentUser):
    a = _q(db, ActionTracker, current_user).filter(ActionTracker.id == action_id).first()
    if not a:
        raise HTTPException(404, "Action not found")
    for k, val in body.model_dump(exclude_unset=True).items():
        setattr(a, k, val)
    db.commit()
    db.refresh(a)
    return ActionTrackerOut.model_validate(a)


@router.delete("/actions/{action_id}", status_code=204)
def delete_action(action_id: int, db: DbSession, current_user: CurrentUser):
    a = _q(db, ActionTracker, current_user).filter(ActionTracker.id == action_id).first()
    if not a:
        raise HTTPException(404, "Action not found")
    a.is_deleted = True
    db.commit()


# ──────────────────────────────────────────────
# PLANT CRUD
# ──────────────────────────────────────────────

@router.get("/plants", response_model=dict)
def list_plants(
    db: DbSession, current_user: CurrentUser,
    page: int = Query(1, ge=1), page_size: int = Query(20, ge=1, le=200),
    search: str = "",
):
    q = _q(db, Plant, current_user)
    if search:
        q = q.filter(Plant.name.ilike(f"%{search}%") | Plant.code.ilike(f"%{search}%"))
    q = q.order_by(Plant.id.desc())
    result = _paginate(q, page, page_size)
    result["items"] = [PlantOut.model_validate(i) for i in result["items"]]
    return result

@router.get("/plants/{plant_id}", response_model=PlantOut)
def get_plant(plant_id: int, db: DbSession, current_user: CurrentUser):
    p = _q(db, Plant, current_user).filter(Plant.id == plant_id).first()
    if not p:
        raise HTTPException(404, "Plant not found")
    return PlantOut.model_validate(p)

@router.post("/plants", response_model=PlantOut, status_code=201)
def create_plant(body: PlantCreate, db: DbSession, current_user: CurrentUser):
    existing = _q(db, Plant, current_user).filter(Plant.code == body.code).first()
    if existing:
        raise HTTPException(409, "Plant code already exists")
    p = Plant(**body.model_dump(), tenant_id=current_user.tenant_id)
    db.add(p); db.commit(); db.refresh(p)
    return PlantOut.model_validate(p)

@router.put("/plants/{plant_id}", response_model=PlantOut)
def update_plant(plant_id: int, body: PlantUpdate, db: DbSession, current_user: CurrentUser):
    p = _q(db, Plant, current_user).filter(Plant.id == plant_id).first()
    if not p:
        raise HTTPException(404, "Plant not found")
    for k, v in body.model_dump(exclude_unset=True).items():
        setattr(p, k, v)
    db.commit(); db.refresh(p)
    return PlantOut.model_validate(p)

@router.delete("/plants/{plant_id}", status_code=204)
def delete_plant(plant_id: int, db: DbSession, current_user: CurrentUser):
    p = _q(db, Plant, current_user).filter(Plant.id == plant_id).first()
    if not p:
        raise HTTPException(404, "Plant not found")
    p.is_deleted = True
    db.commit()

# ──────────────────────────────────────────────
# WAREHOUSE CRUD
# ──────────────────────────────────────────────

@router.get("/warehouses", response_model=dict)
def list_warehouses(
    db: DbSession, current_user: CurrentUser,
    page: int = Query(1, ge=1), page_size: int = Query(20, ge=1, le=200),
    search: str = "",
):
    q = _q(db, Warehouse, current_user)
    if search:
        q = q.filter(Warehouse.name.ilike(f"%{search}%") | Warehouse.code.ilike(f"%{search}%"))
    q = q.order_by(Warehouse.id.desc())
    result = _paginate(q, page, page_size)
    result["items"] = [WarehouseOut.model_validate(i) for i in result["items"]]
    return result

@router.get("/warehouses/{wh_id}", response_model=WarehouseOut)
def get_warehouse(wh_id: int, db: DbSession, current_user: CurrentUser):
    w = _q(db, Warehouse, current_user).filter(Warehouse.id == wh_id).first()
    if not w:
        raise HTTPException(404, "Warehouse not found")
    return WarehouseOut.model_validate(w)

@router.post("/warehouses", response_model=WarehouseOut, status_code=201)
def create_warehouse(body: WarehouseCreate, db: DbSession, current_user: CurrentUser):
    existing = _q(db, Warehouse, current_user).filter(Warehouse.code == body.code).first()
    if existing:
        raise HTTPException(409, "Warehouse code already exists")
    w = Warehouse(**body.model_dump(), tenant_id=current_user.tenant_id)
    db.add(w); db.commit(); db.refresh(w)
    return WarehouseOut.model_validate(w)

@router.put("/warehouses/{wh_id}", response_model=WarehouseOut)
def update_warehouse(wh_id: int, body: WarehouseUpdate, db: DbSession, current_user: CurrentUser):
    w = _q(db, Warehouse, current_user).filter(Warehouse.id == wh_id).first()
    if not w:
        raise HTTPException(404, "Warehouse not found")
    for k, v in body.model_dump(exclude_unset=True).items():
        setattr(w, k, v)
    db.commit(); db.refresh(w)
    return WarehouseOut.model_validate(w)

@router.delete("/warehouses/{wh_id}", status_code=204)
def delete_warehouse(wh_id: int, db: DbSession, current_user: CurrentUser):
    w = _q(db, Warehouse, current_user).filter(Warehouse.id == wh_id).first()
    if not w:
        raise HTTPException(404, "Warehouse not found")
    w.is_deleted = True
    db.commit()

# ──────────────────────────────────────────────
# REGION CRUD
# ──────────────────────────────────────────────

@router.get("/regions", response_model=dict)
def list_regions(
    db: DbSession, current_user: CurrentUser,
    page: int = Query(1, ge=1), page_size: int = Query(20, ge=1, le=200),
    search: str = "",
):
    q = _q(db, Region, current_user)
    if search:
        q = q.filter(Region.name.ilike(f"%{search}%") | Region.code.ilike(f"%{search}%"))
    q = q.order_by(Region.id.desc())
    result = _paginate(q, page, page_size)
    result["items"] = [RegionOut.model_validate(i) for i in result["items"]]
    return result

@router.get("/regions/{region_id}", response_model=RegionOut)
def get_region(region_id: int, db: DbSession, current_user: CurrentUser):
    r = _q(db, Region, current_user).filter(Region.id == region_id).first()
    if not r:
        raise HTTPException(404, "Region not found")
    return RegionOut.model_validate(r)

@router.post("/regions", response_model=RegionOut, status_code=201)
def create_region(body: RegionCreate, db: DbSession, current_user: CurrentUser):
    existing = _q(db, Region, current_user).filter(Region.code == body.code).first()
    if existing:
        raise HTTPException(409, "Region code already exists")
    r = Region(**body.model_dump(), tenant_id=current_user.tenant_id)
    db.add(r); db.commit(); db.refresh(r)
    return RegionOut.model_validate(r)

@router.put("/regions/{region_id}", response_model=RegionOut)
def update_region(region_id: int, body: RegionUpdate, db: DbSession, current_user: CurrentUser):
    r = _q(db, Region, current_user).filter(Region.id == region_id).first()
    if not r:
        raise HTTPException(404, "Region not found")
    for k, v in body.model_dump(exclude_unset=True).items():
        setattr(r, k, v)
    db.commit(); db.refresh(r)
    return RegionOut.model_validate(r)

@router.delete("/regions/{region_id}", status_code=204)
def delete_region(region_id: int, db: DbSession, current_user: CurrentUser):
    r = _q(db, Region, current_user).filter(Region.id == region_id).first()
    if not r:
        raise HTTPException(404, "Region not found")
    r.is_deleted = True
    db.commit()

# ──────────────────────────────────────────────
# BUSINESS UNIT CRUD
# ──────────────────────────────────────────────

@router.get("/business-units", response_model=dict)
def list_business_units(
    db: DbSession, current_user: CurrentUser,
    page: int = Query(1, ge=1), page_size: int = Query(20, ge=1, le=200),
    search: str = "",
):
    q = _q(db, BusinessUnit, current_user)
    if search:
        q = q.filter(BusinessUnit.name.ilike(f"%{search}%") | BusinessUnit.code.ilike(f"%{search}%"))
    q = q.order_by(BusinessUnit.id.desc())
    result = _paginate(q, page, page_size)
    result["items"] = [BusinessUnitOut.model_validate(i) for i in result["items"]]
    return result

@router.get("/business-units/{bu_id}", response_model=BusinessUnitOut)
def get_business_unit(bu_id: int, db: DbSession, current_user: CurrentUser):
    b = _q(db, BusinessUnit, current_user).filter(BusinessUnit.id == bu_id).first()
    if not b:
        raise HTTPException(404, "Business unit not found")
    return BusinessUnitOut.model_validate(b)

@router.post("/business-units", response_model=BusinessUnitOut, status_code=201)
def create_business_unit(body: BusinessUnitCreate, db: DbSession, current_user: CurrentUser):
    existing = _q(db, BusinessUnit, current_user).filter(BusinessUnit.code == body.code).first()
    if existing:
        raise HTTPException(409, "Business unit code already exists")
    b = BusinessUnit(**body.model_dump(), tenant_id=current_user.tenant_id)
    db.add(b); db.commit(); db.refresh(b)
    return BusinessUnitOut.model_validate(b)

@router.put("/business-units/{bu_id}", response_model=BusinessUnitOut)
def update_business_unit(bu_id: int, body: BusinessUnitUpdate, db: DbSession, current_user: CurrentUser):
    b = _q(db, BusinessUnit, current_user).filter(BusinessUnit.id == bu_id).first()
    if not b:
        raise HTTPException(404, "Business unit not found")
    for k, v in body.model_dump(exclude_unset=True).items():
        setattr(b, k, v)
    db.commit(); db.refresh(b)
    return BusinessUnitOut.model_validate(b)

@router.delete("/business-units/{bu_id}", status_code=204)
def delete_business_unit(bu_id: int, db: DbSession, current_user: CurrentUser):
    b = _q(db, BusinessUnit, current_user).filter(BusinessUnit.id == bu_id).first()
    if not b:
        raise HTTPException(404, "Business unit not found")
    b.is_deleted = True
    db.commit()

# ──────────────────────────────────────────────
# RISK & CONTROL MATRIX CRUD
# ──────────────────────────────────────────────

@router.get("/risk-controls", response_model=dict)
def list_risk_controls(
    db: DbSession, current_user: CurrentUser,
    page: int = Query(1, ge=1), page_size: int = Query(20, ge=1, le=200),
    search: str = "",
    risk_category: str = "",
):
    q = _q(db, RiskControl, current_user)
    if search:
        q = q.filter(RiskControl.risk_description.ilike(f"%{search}%") | RiskControl.risk_code.ilike(f"%{search}%"))
    if risk_category:
        q = q.filter(RiskControl.risk_category == risk_category)
    q = q.order_by(RiskControl.id.desc())
    result = _paginate(q, page, page_size)
    result["items"] = [RiskControlOut.model_validate(i) for i in result["items"]]
    return result

@router.get("/risk-controls/{rc_id}", response_model=RiskControlOut)
def get_risk_control(rc_id: int, db: DbSession, current_user: CurrentUser):
    rc = _q(db, RiskControl, current_user).filter(RiskControl.id == rc_id).first()
    if not rc:
        raise HTTPException(404, "Risk control not found")
    return RiskControlOut.model_validate(rc)

@router.post("/risk-controls", response_model=RiskControlOut, status_code=201)
def create_risk_control(body: RiskControlCreate, db: DbSession, current_user: CurrentUser):
    existing = _q(db, RiskControl, current_user).filter(RiskControl.risk_code == body.risk_code).first()
    if existing:
        raise HTTPException(409, "Risk code already exists")
    rc = RiskControl(**body.model_dump(), tenant_id=current_user.tenant_id)
    db.add(rc); db.commit(); db.refresh(rc)
    return RiskControlOut.model_validate(rc)

@router.put("/risk-controls/{rc_id}", response_model=RiskControlOut)
def update_risk_control(rc_id: int, body: RiskControlUpdate, db: DbSession, current_user: CurrentUser):
    rc = _q(db, RiskControl, current_user).filter(RiskControl.id == rc_id).first()
    if not rc:
        raise HTTPException(404, "Risk control not found")
    for k, v in body.model_dump(exclude_unset=True).items():
        setattr(rc, k, v)
    db.commit(); db.refresh(rc)
    return RiskControlOut.model_validate(rc)

@router.delete("/risk-controls/{rc_id}", status_code=204)
def delete_risk_control(rc_id: int, db: DbSession, current_user: CurrentUser):
    rc = _q(db, RiskControl, current_user).filter(RiskControl.id == rc_id).first()
    if not rc:
        raise HTTPException(404, "Risk control not found")
    rc.is_deleted = True
    db.commit()

# ──────────────────────────────────────────────
# TEST RULE LIBRARY CRUD
# ──────────────────────────────────────────────

@router.get("/test-rules", response_model=dict)
def list_test_rules(
    db: DbSession, current_user: CurrentUser,
    page: int = Query(1, ge=1), page_size: int = Query(20, ge=1, le=200),
    search: str = "",
    category: str = "",
):
    q = _q(db, TestRule, current_user)
    if search:
        q = q.filter(TestRule.rule_name.ilike(f"%{search}%") | TestRule.rule_code.ilike(f"%{search}%"))
    if category:
        q = q.filter(TestRule.category == category)
    q = q.order_by(TestRule.id.desc())
    result = _paginate(q, page, page_size)
    result["items"] = [TestRuleOut.model_validate(i) for i in result["items"]]
    return result

@router.get("/test-rules/{rule_id}", response_model=TestRuleOut)
def get_test_rule(rule_id: int, db: DbSession, current_user: CurrentUser):
    tr = _q(db, TestRule, current_user).filter(TestRule.id == rule_id).first()
    if not tr:
        raise HTTPException(404, "Test rule not found")
    return TestRuleOut.model_validate(tr)

@router.post("/test-rules", response_model=TestRuleOut, status_code=201)
def create_test_rule(body: TestRuleCreate, db: DbSession, current_user: CurrentUser):
    existing = _q(db, TestRule, current_user).filter(TestRule.rule_code == body.rule_code).first()
    if existing:
        raise HTTPException(409, "Rule code already exists")
    tr = TestRule(**body.model_dump(), tenant_id=current_user.tenant_id)
    db.add(tr); db.commit(); db.refresh(tr)
    return TestRuleOut.model_validate(tr)

@router.put("/test-rules/{rule_id}", response_model=TestRuleOut)
def update_test_rule(rule_id: int, body: TestRuleUpdate, db: DbSession, current_user: CurrentUser):
    tr = _q(db, TestRule, current_user).filter(TestRule.id == rule_id).first()
    if not tr:
        raise HTTPException(404, "Test rule not found")
    for k, v in body.model_dump(exclude_unset=True).items():
        setattr(tr, k, v)
    db.commit(); db.refresh(tr)
    return TestRuleOut.model_validate(tr)

@router.delete("/test-rules/{rule_id}", status_code=204)
def delete_test_rule(rule_id: int, db: DbSession, current_user: CurrentUser):
    tr = _q(db, TestRule, current_user).filter(TestRule.id == rule_id).first()
    if not tr:
        raise HTTPException(404, "Test rule not found")
    tr.is_deleted = True
    db.commit()

# ──────────────────────────────────────────────
# DATA SOURCE CRUD
# ──────────────────────────────────────────────

@router.get("/data-sources", response_model=dict)
def list_data_sources(
    db: DbSession, current_user: CurrentUser,
    page: int = Query(1, ge=1), page_size: int = Query(20, ge=1, le=200),
    search: str = "",
    source_type: str = "",
):
    q = _q(db, DataSource, current_user)
    if search:
        q = q.filter(DataSource.source_name.ilike(f"%{search}%") | DataSource.source_code.ilike(f"%{search}%"))
    if source_type:
        q = q.filter(DataSource.source_type == source_type)
    q = q.order_by(DataSource.id.desc())
    result = _paginate(q, page, page_size)
    result["items"] = [DataSourceOut.model_validate(i) for i in result["items"]]
    return result

@router.get("/data-sources/{ds_id}", response_model=DataSourceOut)
def get_data_source(ds_id: int, db: DbSession, current_user: CurrentUser):
    ds = _q(db, DataSource, current_user).filter(DataSource.id == ds_id).first()
    if not ds:
        raise HTTPException(404, "Data source not found")
    return DataSourceOut.model_validate(ds)

@router.post("/data-sources", response_model=DataSourceOut, status_code=201)
def create_data_source(body: DataSourceCreate, db: DbSession, current_user: CurrentUser):
    existing = _q(db, DataSource, current_user).filter(DataSource.source_code == body.source_code).first()
    if existing:
        raise HTTPException(409, "Source code already exists")
    ds = DataSource(**body.model_dump(), tenant_id=current_user.tenant_id)
    db.add(ds); db.commit(); db.refresh(ds)
    return DataSourceOut.model_validate(ds)

@router.put("/data-sources/{ds_id}", response_model=DataSourceOut)
def update_data_source(ds_id: int, body: DataSourceUpdate, db: DbSession, current_user: CurrentUser):
    ds = _q(db, DataSource, current_user).filter(DataSource.id == ds_id).first()
    if not ds:
        raise HTTPException(404, "Data source not found")
    for k, v in body.model_dump(exclude_unset=True).items():
        setattr(ds, k, v)
    db.commit(); db.refresh(ds)
    return DataSourceOut.model_validate(ds)

@router.delete("/data-sources/{ds_id}", status_code=204)
def delete_data_source(ds_id: int, db: DbSession, current_user: CurrentUser):
    ds = _q(db, DataSource, current_user).filter(DataSource.id == ds_id).first()
    if not ds:
        raise HTTPException(404, "Data source not found")
    ds.is_deleted = True
    db.commit()

# ──────────────────────────────────────────────
# SAMPLING RECORDS CRUD
# ──────────────────────────────────────────────

@router.get("/sampling", response_model=dict)
def list_sampling_records(
    db: DbSession, current_user: CurrentUser,
    page: int = Query(1, ge=1), page_size: int = Query(20, ge=1, le=200),
    search: str = "",
    sampling_method: str = "",
):
    q = _q(db, SamplingRecord, current_user)
    if search:
        q = q.filter(SamplingRecord.sample_code.ilike(f"%{search}%"))
    if sampling_method:
        q = q.filter(SamplingRecord.sampling_method == sampling_method)
    q = q.order_by(SamplingRecord.id.desc())
    result = _paginate(q, page, page_size)
    result["items"] = [SamplingRecordOut.model_validate(i) for i in result["items"]]
    return result

@router.get("/sampling/{sample_id}", response_model=SamplingRecordOut)
def get_sampling_record(sample_id: int, db: DbSession, current_user: CurrentUser):
    sr = _q(db, SamplingRecord, current_user).filter(SamplingRecord.id == sample_id).first()
    if not sr:
        raise HTTPException(404, "Sampling record not found")
    return SamplingRecordOut.model_validate(sr)

@router.post("/sampling", response_model=SamplingRecordOut, status_code=201)
def create_sampling_record(body: SamplingRecordCreate, db: DbSession, current_user: CurrentUser):
    sr = SamplingRecord(**body.model_dump(), tenant_id=current_user.tenant_id)
    db.add(sr); db.commit(); db.refresh(sr)
    return SamplingRecordOut.model_validate(sr)

@router.put("/sampling/{sample_id}", response_model=SamplingRecordOut)
def update_sampling_record(sample_id: int, body: SamplingRecordUpdate, db: DbSession, current_user: CurrentUser):
    sr = _q(db, SamplingRecord, current_user).filter(SamplingRecord.id == sample_id).first()
    if not sr:
        raise HTTPException(404, "Sampling record not found")
    for k, v in body.model_dump(exclude_unset=True).items():
        setattr(sr, k, v)
    db.commit(); db.refresh(sr)
    return SamplingRecordOut.model_validate(sr)

@router.delete("/sampling/{sample_id}", status_code=204)
def delete_sampling_record(sample_id: int, db: DbSession, current_user: CurrentUser):
    sr = _q(db, SamplingRecord, current_user).filter(SamplingRecord.id == sample_id).first()
    if not sr:
        raise HTTPException(404, "Sampling record not found")
    sr.is_deleted = True
    db.commit()

@router.post("/sampling/generate", response_model=SamplingRecordOut, status_code=201)
def generate_sample(body: SamplingRecordCreate, db: DbSession, current_user: CurrentUser):
    """Generate a statistical sample from a specified population."""
    import json, random
    pop_table = body.population_table
    pop_count = body.population_count
    sample_size = body.sample_size
    method = body.sampling_method

    population = list(range(1, pop_count + 1))
    selected = []
    if method == "Random":
        selected = random.sample(population, min(sample_size, pop_count))
    elif method == "Systematic":
        step = max(1, pop_count // sample_size)
        start = random.randint(1, step)
        selected = list(range(start, pop_count + 1, step))[:sample_size]
    elif method == "Stratified":
        strata = pop_count // 5
        per_stratum = max(1, sample_size // 5)
        for s in range(5):
            stratum = list(range(s * strata + 1, min((s + 1) * strata, pop_count) + 1))
            selected.extend(random.sample(stratum, min(per_stratum, len(stratum))))
    else:
        selected = sorted(population[:sample_size])

    sr = SamplingRecord(
        sample_code=body.sample_code,
        population_table=pop_table,
        population_count=pop_count,
        sample_size=len(selected),
        sampling_method=method,
        confidence_level=body.confidence_level,
        margin_of_error=body.margin_of_error,
        sample_data=json.dumps(selected),
        status="Generated",
        created_by=body.created_by or "system",
        tenant_id=current_user.tenant_id,
    )
    db.add(sr); db.commit(); db.refresh(sr)
    return SamplingRecordOut.model_validate(sr)

# ──────────────────────────────────────────────
# EXCEPTION QUEUE CRUD
# ──────────────────────────────────────────────

@router.get("/exceptions", response_model=dict)
def list_exceptions(
    db: DbSession, current_user: CurrentUser,
    page: int = Query(1, ge=1), page_size: int = Query(20, ge=1, le=200),
    search: str = "",
    severity: str = "",
    status: str = "",
):
    q = _q(db, ExceptionItem, current_user)
    if search:
        q = q.filter(ExceptionItem.title.ilike(f"%{search}%") | ExceptionItem.exception_code.ilike(f"%{search}%"))
    if severity:
        q = q.filter(ExceptionItem.severity == severity)
    if status:
        q = q.filter(ExceptionItem.status == status)
    q = q.order_by(ExceptionItem.id.desc())
    result = _paginate(q, page, page_size)
    result["items"] = [ExceptionItemOut.model_validate(i) for i in result["items"]]
    return result

@router.get("/exceptions/{exc_id}", response_model=ExceptionItemOut)
def get_exception(exc_id: int, db: DbSession, current_user: CurrentUser):
    e = _q(db, ExceptionItem, current_user).filter(ExceptionItem.id == exc_id).first()
    if not e:
        raise HTTPException(404, "Exception not found")
    return ExceptionItemOut.model_validate(e)

@router.post("/exceptions", response_model=ExceptionItemOut, status_code=201)
def create_exception(body: ExceptionItemCreate, db: DbSession, current_user: CurrentUser):
    e = ExceptionItem(**body.model_dump(), tenant_id=current_user.tenant_id)
    db.add(e); db.commit(); db.refresh(e)
    return ExceptionItemOut.model_validate(e)

@router.put("/exceptions/{exc_id}", response_model=ExceptionItemOut)
def update_exception(exc_id: int, body: ExceptionItemUpdate, db: DbSession, current_user: CurrentUser):
    e = _q(db, ExceptionItem, current_user).filter(ExceptionItem.id == exc_id).first()
    if not e:
        raise HTTPException(404, "Exception not found")
    for k, v in body.model_dump(exclude_unset=True).items():
        setattr(e, k, v)
    db.commit(); db.refresh(e)
    return ExceptionItemOut.model_validate(e)

@router.delete("/exceptions/{exc_id}", status_code=204)
def delete_exception(exc_id: int, db: DbSession, current_user: CurrentUser):
    e = _q(db, ExceptionItem, current_user).filter(ExceptionItem.id == exc_id).first()
    if not e:
        raise HTTPException(404, "Exception not found")
    e.is_deleted = True
    db.commit()

# ──────────────────────────────────────────────
# WORKING PAPERS CRUD
# ──────────────────────────────────────────────

@router.get("/working-papers", response_model=dict)
def list_working_papers(
    db: DbSession, current_user: CurrentUser,
    page: int = Query(1, ge=1), page_size: int = Query(20, ge=1, le=200),
    search: str = "",
    status: str = "",
):
    q = _q(db, WorkingPaper, current_user)
    if search:
        q = q.filter(WorkingPaper.title.ilike(f"%{search}%") | WorkingPaper.wp_code.ilike(f"%{search}%"))
    if status:
        q = q.filter(WorkingPaper.status == status)
    q = q.order_by(WorkingPaper.id.desc())
    result = _paginate(q, page, page_size)
    result["items"] = [WorkingPaperOut.model_validate(i) for i in result["items"]]
    return result

@router.get("/working-papers/{wp_id}", response_model=WorkingPaperOut)
def get_working_paper(wp_id: int, db: DbSession, current_user: CurrentUser):
    wp = _q(db, WorkingPaper, current_user).filter(WorkingPaper.id == wp_id).first()
    if not wp:
        raise HTTPException(404, "Working paper not found")
    return WorkingPaperOut.model_validate(wp)

@router.post("/working-papers", response_model=WorkingPaperOut, status_code=201)
def create_working_paper(body: WorkingPaperCreate, db: DbSession, current_user: CurrentUser):
    wp = WorkingPaper(**body.model_dump(), tenant_id=current_user.tenant_id)
    db.add(wp); db.commit(); db.refresh(wp)
    return WorkingPaperOut.model_validate(wp)

@router.put("/working-papers/{wp_id}", response_model=WorkingPaperOut)
def update_working_paper(wp_id: int, body: WorkingPaperUpdate, db: DbSession, current_user: CurrentUser):
    wp = _q(db, WorkingPaper, current_user).filter(WorkingPaper.id == wp_id).first()
    if not wp:
        raise HTTPException(404, "Working paper not found")
    for k, v in body.model_dump(exclude_unset=True).items():
        setattr(wp, k, v)
    db.commit(); db.refresh(wp)
    return WorkingPaperOut.model_validate(wp)

@router.delete("/working-papers/{wp_id}", status_code=204)
def delete_working_paper(wp_id: int, db: DbSession, current_user: CurrentUser):
    wp = _q(db, WorkingPaper, current_user).filter(WorkingPaper.id == wp_id).first()
    if not wp:
        raise HTTPException(404, "Working paper not found")
    wp.is_deleted = True
    db.commit()

# ──────────────────────────────────────────────
# SEED ENDPOINT
# ──────────────────────────────────────────────

@router.post("/seed", status_code=201)
def seed_module_data(db: DbSession, current_user: CurrentUser):
    from .seed import seed_data as run_seed
    run_seed(db, current_user.tenant_id)
    return {"message": "Seed data created successfully"}


# ──────────────────────────────────────────────
# DASHBOARD
# ──────────────────────────────────────────────

@router.get("/dashboard")
def dashboard_summary(db: DbSession, current_user: CurrentUser):
    return svc.get_dashboard_summary(db, current_user)


@router.get("/dashboard/trends")
def dashboard_trends(db: DbSession, current_user: CurrentUser):
    return an.freight_cost_per_unit_trend(db, current_user)


# ──────────────────────────────────────────────
# ANALYTICS ENDPOINTS
# ──────────────────────────────────────────────

@router.get("/analytics/rate-compliance")
def analytics_rate_compliance(db: DbSession, current_user: CurrentUser):
    return an.rate_compliance_analysis(db, current_user)


@router.get("/analytics/weight-variance")
def analytics_weight_variance(db: DbSession, current_user: CurrentUser):
    return an.weight_variance_analysis(db, current_user)


@router.get("/analytics/route-distance")
def analytics_route_distance(db: DbSession, current_user: CurrentUser):
    return an.route_distance_analytics(db, current_user)


@router.get("/analytics/detention")
def analytics_detention(db: DbSession, current_user: CurrentUser):
    return an.detention_demurrage_analysis(db, current_user)


@router.get("/analytics/carrier-performance")
def analytics_carrier_performance(db: DbSession, current_user: CurrentUser):
    return an.carrier_performance_analysis(db, current_user)


@router.get("/analytics/duplicate-billing")
def analytics_duplicate_billing(db: DbSession, current_user: CurrentUser):
    return an.duplicate_billing_analysis(db, current_user)


@router.get("/analytics/multimodal-cost")
def analytics_multimodal_cost(db: DbSession, current_user: CurrentUser, shipment_id: int = 0):
    return an.multimodal_cost_comparison(db, current_user, shipment_id or None)


@router.get("/analytics/fuel-surcharge")
def analytics_fuel_surcharge(db: DbSession, current_user: CurrentUser):
    return an.fuel_surcharge_validation(db, current_user)


@router.get("/analytics/empty-return")
def analytics_empty_return(db: DbSession, current_user: CurrentUser):
    return an.empty_return_analysis(db, current_user)


@router.get("/analytics/lr-pod-match")
def analytics_lr_pod_match(db: DbSession, current_user: CurrentUser):
    return an.lr_pod_match_analysis(db, current_user)


@router.get("/analytics/provision-accuracy")
def analytics_provision_accuracy(db: DbSession, current_user: CurrentUser):
    return an.freight_provision_accuracy(db, current_user)


@router.get("/analytics/transit-sla")
def analytics_transit_sla(db: DbSession, current_user: CurrentUser):
    return an.transit_time_sla_analysis(db, current_user)


@router.get("/analytics/damage-claims")
def analytics_damage_claims(db: DbSession, current_user: CurrentUser):
    return an.damage_shortage_claims(db, current_user)


@router.get("/analytics/vehicle-placement")
def analytics_vehicle_placement(db: DbSession, current_user: CurrentUser):
    return an.vehicle_placement_efficiency(db, current_user)


@router.get("/analytics/cost-trends")
def analytics_cost_trends(db: DbSession, current_user: CurrentUser):
    return an.freight_cost_per_unit_trend(db, current_user)


@router.get("/analytics/risk-score")
def analytics_risk_score(db: DbSession, current_user: CurrentUser):
    summary = svc.get_dashboard_summary(db, current_user)
    return {"risk_score": summary["risk_score"], "factors": summary}


# ──────────────────────────────────────────────
# EXPORT ENDPOINTS
# ──────────────────────────────────────────────

@router.get("/export/shipments/csv")
def export_shipments_csv(db: DbSession, current_user: CurrentUser):
    shipments = _q(db, FreightShipment, current_user).order_by(FreightShipment.shipment_date.desc()).all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "ID", "Shipment#", "LR#", "Date", "Origin", "Destination",
        "Mode", "Actual Wt", "Charged Wt", "Contract Rate", "Billed Rate",
        "Freight", "Fuel Surcharge", "Total", "Status"
    ])
    for s in shipments:
        writer.writerow([
            s.id, s.shipment_number, s.lr_number, s.shipment_date,
            s.origin, s.destination, s.mode, s.actual_weight_kg,
            s.charged_weight_kg, s.contract_rate, s.billed_rate,
            s.freight_amount, s.fuel_surcharge, s.total_amount, s.status,
        ])
    return Response(content=output.getvalue(), media_type="text/csv",
                    headers={"Content-Disposition": "attachment; filename=shipments.csv"})


@router.get("/export/invoices/csv")
def export_invoices_csv(db: DbSession, current_user: CurrentUser):
    invoices = _q(db, FreightInvoice, current_user).order_by(FreightInvoice.invoice_date.desc()).all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "ID", "Invoice#", "Date", "Due Date", "Billed", "Approved",
        "Difference", "Fuel Surcharge", "Total", "Status", "Payment Status",
    ])
    for i in invoices:
        writer.writerow([
            i.id, i.invoice_number, i.invoice_date, i.due_date,
            i.billed_amount, i.approved_amount, i.difference_amount,
            i.fuel_surcharge_billed, i.total_amount, i.status, i.payment_status,
        ])
    return Response(content=output.getvalue(), media_type="text/csv",
                    headers={"Content-Disposition": "attachment; filename=invoices.csv"})


@router.get("/export/findings/csv")
def export_findings_csv(db: DbSession, current_user: CurrentUser):
    findings = _q(db, Finding, current_user).order_by(Finding.id.desc()).all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "Finding#", "Title", "Category", "Severity", "Status", "Financial Impact"])
    for f in findings:
        writer.writerow([f.id, f.finding_number, f.title, f.category, f.severity, f.status, f.financial_impact])
    return Response(content=output.getvalue(), media_type="text/csv",
                    headers={"Content-Disposition": "attachment; filename=findings.csv"})


@router.get("/export/carriers/csv")
def export_carriers_csv(db: DbSession, current_user: CurrentUser):
    carriers = _q(db, Carrier, current_user).order_by(Carrier.name).all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "Code", "Name", "Type", "Status", "Score", "On-Time%", "Damage%", "Claim%", "Delay%"])
    for c in carriers:
        writer.writerow([
            c.id, c.code, c.name, c.carrier_type, c.status,
            c.performance_score, c.on_time_percentage, c.damage_percentage,
            c.claim_percentage, c.delay_percentage,
        ])
    return Response(content=output.getvalue(), media_type="text/csv",
                    headers={"Content-Disposition": "attachment; filename=carriers.csv"})


@router.get("/export/claims/csv")
def export_claims_csv(db: DbSession, current_user: CurrentUser):
    claims = _q(db, Claim, current_user).order_by(Claim.claim_date.desc()).all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "Claim#", "Type", "Date", "Value", "Recovered", "Pending", "Rejected", "Status"])
    for c in claims:
        writer.writerow([
            c.id, c.claim_number, c.claim_type, c.claim_date,
            c.claim_value, c.recovered_amount, c.pending_amount,
            c.rejected_amount, c.status,
        ])
    return Response(content=output.getvalue(), media_type="text/csv",
                    headers={"Content-Disposition": "attachment; filename=claims.csv"})


@router.get("/export/shipments/xlsx")
def export_shipments_xlsx(db: DbSession, current_user: CurrentUser):
    return _export_xlsx(
        db, FreightShipment, current_user,
        ["ID", "Shipment#", "LR#", "Date", "Origin", "Destination", "Mode", "Actual Wt", "Charged Wt", "Freight", "Status"],
        lambda s: [s.id, s.shipment_number, s.lr_number, str(s.shipment_date), s.origin, s.destination, s.mode, s.actual_weight_kg, s.charged_weight_kg, s.freight_amount, s.status],
        "shipments.xlsx"
    )


@router.get("/export/invoices/xlsx")
def export_invoices_xlsx(db: DbSession, current_user: CurrentUser):
    return _export_xlsx(
        db, FreightInvoice, current_user,
        ["ID", "Invoice#", "Date", "Due Date", "Billed", "Approved", "Difference", "Total", "Status"],
        lambda i: [i.id, i.invoice_number, str(i.invoice_date), str(i.due_date), i.billed_amount, i.approved_amount, i.difference_amount, i.total_amount, i.status],
        "invoices.xlsx"
    )


def _export_xlsx(db, model, current_user, headers, row_fn, filename):
    items = _q(db, model, current_user).order_by(model.id.desc()).all()
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Export"
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
        for row_idx, item in enumerate(items, 2):
            for col_idx, val in enumerate(row_fn(item), 1):
                ws.cell(row=row_idx, column=col_idx, value=val)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return Response(content=output.read(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": f"attachment; filename={filename}"})
    except ImportError:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for item in items:
            writer.writerow(row_fn(item))
        return Response(content=output.getvalue(), media_type="text/csv",
                        headers={"Content-Disposition": f"attachment; filename={filename.replace('.xlsx', '.csv')}"})
